######################################################################################################
# Background docs: https://cloud.gov/docs/compliance/auditing-activity/.
# There is a ton of duplicated code in here that I plan to refactor. Basically, v3 of the API
# seems much more consistent than v2 which will allow us to consolidate code sections.
######################################################################################################

import subprocess
import sys
import getopt
import os
import json
import hashlib
import re
from datetime import datetime, time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

start_path = os.getcwd()
data_folder_path = start_path
cf_guid_dict = dict()

######################################################################################################
# Generic command line functions.
######################################################################################################


def run_cmd_suppress_output(cmd, args):
    subprocess.call([cmd, args], stdout=open(
        os.devnull, "w"), stderr=subprocess.STDOUT)
    return


def run_cli_cmd(cmd, filename):
    p1 = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    output = p1.communicate()[0]
    f = open(filename + ".txt", "w")
    print(output.decode(), file=f)
    f.close()
    return


def run_api_cmd(cmd, filename):
    p1 = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    output = p1.communicate()[0]
    f = open(filename + ".json", "w")
    print(output.decode(), file=f)
    f.close()
    return


def run_api_cmd_rtn_json(cmd):
    p1 = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    output = p1.communicate()[0]
    return output.decode()


def run_api_cmd_and_hash_output_txt(cmd, hash_statement, json_attrib, filename):
    p1 = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    output = p1.communicate()[0]
    json_data = json.loads(output)

    text = ""
    if "CF-NotAuthorized" in output.decode():
        text = "Not authorized to perform action around this action."
        hash_statement = text
        print(text + " " + '/'.join(cmd))
    elif json_attrib != None and json_attrib not in output.decode():
        text = "There is no " + json_attrib + " in the results of this action."
        hash_statement = text
    elif json_attrib == None:
        text = json_data
    else:
        text = json_data[json_attrib]

    hash_object = hashlib.md5(json.dumps(text).encode('utf-8'))
    f = open(filename + ".txt", "w")
    print(hash_statement + ": " + hash_object.hexdigest(), file=f)
    f.close()
    return


def remove_prefix(s, prefix):
    return s[len(prefix):] if s.startswith(prefix) else s


def getValueUsingGUID(data, guidToCheck, defaultRtnvalue):

    for record in data:
        if record["guid"] == guidToCheck:
            return record["name"]
    return defaultRtnvalue


def GetKey(val):
    for key, value in cf_guid_dict.items():
        if val == value:
            return key
    return None


######################################################################################################
######################################################################################################


def write_events_to_sheet(sheet0, data, audit_event_types, event_start_dt, event_end_dt):
    HeaderColorFill = PatternFill(fgColor="C0C0C0", fill_type="solid")

    sheet0["A1"] = "Timestamp"
    sheet0["A1"].fill = HeaderColorFill
    sheet0["A1"].font = Font(bold=True)

    sheet0["B1"] = "Organization"
    sheet0["B1"].fill = HeaderColorFill
    sheet0["B1"].font = Font(bold=True)

    sheet0["C1"] = "Space"
    sheet0["C1"].fill = HeaderColorFill
    sheet0["C1"].font = Font(bold=True)

    sheet0["D1"] = "Actor Type"
    sheet0["D1"].fill = HeaderColorFill
    sheet0["D1"].font = Font(bold=True)

    sheet0["E1"] = "Actor"
    sheet0["E1"].fill = HeaderColorFill
    sheet0["E1"].font = Font(bold=True)

    sheet0["F1"] = "Action Type"
    sheet0["F1"].fill = HeaderColorFill
    sheet0["F1"].font = Font(bold=True)

    sheet0["G1"] = "Target Type"
    sheet0["G1"].fill = HeaderColorFill
    sheet0["G1"].font = Font(bold=True)

    sheet0["H1"] = "Target User"
    sheet0["H1"].fill = HeaderColorFill
    sheet0["H1"].font = Font(bold=True)

    sheet0["I1"] = "Data"
    sheet0["I1"].fill = HeaderColorFill
    sheet0["I1"].font = Font(bold=True)

    sheet0["J1"] = "Event Link"
    sheet0["J1"].fill = HeaderColorFill
    sheet0["J1"].font = Font(bold=True)

    sheet0["K1"] = "Source"
    sheet0["K1"].fill = HeaderColorFill
    sheet0["K1"].font = Font(bold=True)

    done = False
    ExcelRowRef = 1

    for d in data["resources"]:
        if audit_event_types == None or audit_event_types == "" or (audit_event_types != None and d["type"] in audit_event_types):
            event_dt = datetime.strptime(d["created_at"], '%Y-%m-%dT%H:%M:%SZ')

            if event_start_dt != None and event_end_dt == None:
                if event_dt < event_start_dt:
                    continue
            elif  event_start_dt == None and event_end_dt != None:
                if event_dt > event_end_dt:
                    continue
            elif event_start_dt != None and event_end_dt != None: #between
                if event_dt < event_start_dt or event_dt > event_end_dt:
                    continue

            ExcelRowRef += 1

            sheet0["A" + str(ExcelRowRef)] = d["created_at"]
            sheet0["B" + str(ExcelRowRef)] = getValueUsingGUID(data["org"],
                                                               d["organization"]["guid"], "MISSING ORG NAME")

            space_guid = ""
            try:
                space_guid = d["space"]["guid"]
            except:
                space_guid = ""

            sheet0["C" + str(ExcelRowRef)] = getValueUsingGUID(data["space"],
                                                               space_guid, "MISSING SPACE NAME")
            sheet0["D" + str(ExcelRowRef)] = d["actor"]["type"]
            sheet0["E" + str(ExcelRowRef)] = d["actor"]["name"]
            sheet0["F" + str(ExcelRowRef)] = d["type"]
            sheet0["G" + str(ExcelRowRef)] = d["target"]["type"]
            sheet0["H" + str(ExcelRowRef)] = d["target"]["name"]
            sheet0["I" + str(ExcelRowRef)] = json.dumps(d["data"])
            sheet0["J" + str(ExcelRowRef)] = "cf7 curl "
            sheet0["K" + str(ExcelRowRef)] = json.dumps(d)

    for col in sheet0.columns:
        max_lenght = 0
        col_name = re.findall('\w\d', str(col[0]))
        col_name = col_name[0]
        col_name = re.findall('\w', str(col_name))[0]
        for cell in col:
            try:
                if len(str(cell.value)) > max_lenght:
                    max_lenght = len(cell.value)
            except:
                pass
        adjusted_width = (max_lenght+2)
        sheet0.column_dimensions[col_name].width = adjusted_width

    return


######################################################################################################
######################################################################################################


def archive_all_audit_events_by_type(data, output_file, event_start_dt, event_end_dt):

    Excelworkbook = Workbook()

    print("Generating audit event report...")
    print("All events...")
    sheet0 = Excelworkbook.active
    sheet0.title = "All Events"
    audit_event_types = ""
    write_events_to_sheet(sheet0, data, audit_event_types, event_start_dt, event_end_dt)

    print("User Access Changes...")
    sheet1 = Excelworkbook.create_sheet("User Access Changes", 1)
    audit_event_types = "audit.user.space_developer_add,audit.user.space_developer_remove,audit.user.space_auditor_add,audit.user.space_auditor_remove,audit.user.space_manager_add,audit.user.space_manager_remove"
    write_events_to_sheet(sheet1, data, audit_event_types, event_start_dt, event_end_dt)

    print("Route Changes...")
    sheet2 = Excelworkbook.create_sheet("Route Changes", 2)
    audit_event_types = "audit.route.create,audit.route.delete-request,audit.route.update"
    write_events_to_sheet(sheet2, data, audit_event_types, event_start_dt, event_end_dt)

    print("Service Instance Events...")
    sheet3 = Excelworkbook.create_sheet("Service Instance Events", 3)
    audit_event_types = "audit.service_instance.create,audit.service_instance.bind_route,audit.service_instance.update,audit.service_instance.unbind_route,audit.service_instance.delete"
    write_events_to_sheet(sheet3, data, audit_event_types, event_start_dt, event_end_dt)

    print("Service Bindings...")
    sheet4 = Excelworkbook.create_sheet("Service Bindings", 4)
    audit_event_types = "audit.service_binding.create,service_instance.bind_route,audit.service_instance.unbind_route"
    write_events_to_sheet(sheet4, data, audit_event_types, event_start_dt, event_end_dt)

    print("Service Events...")
    sheet5 = Excelworkbook.create_sheet("Service Events", 5)
    audit_event_types = "audit.service.create,audit.service.delete,audit.service.update,audit.service_binding.create,audit.service_binding.delete,service_instance.bind_route,audit.service_instance.create,audit.service_instance.delete,audit.service_instance.unbind_route,audit.service_instance.update"
    write_events_to_sheet(sheet5, data, audit_event_types, event_start_dt, event_end_dt)

    print()

    Excelworkbook.save(filename=output_file)
    return


######################################################################################################
######################################################################################################


def start(input_file, output_file, event_start_dt, event_end_dt):

    with open(input_file, 'r') as f:
        text = f.read()

    data = json.loads(text)

    archive_all_audit_events_by_type(data, output_file, event_start_dt, event_end_dt)

    return

######################################################################################################
######################################################################################################


full_cmd_arguments = sys.argv
argument_list = full_cmd_arguments[1:]
short_options = "i:o:s:e:"
long_options = ["input_file=", "output_file=", "event_start_dt=", "event_end_dt="]
event_start_dt = None
event_end_dt = None

try:
    arguments, values = getopt.getopt(
        argument_list, short_options, long_options)
except getopt.error as err:
    print(str(err))
    sys.exit(-1)

o_name = ""
file_name = ""
for current_argument, current_value in arguments:
    if current_argument in ("-i", "--input_file"):
        input_file = current_value
    if current_argument in ("-o", "--output_file"):
        output_file = current_value
    if current_argument in ("-s", "--event_start_dt"):
        event_start_dt = current_value
    if current_argument in ("-e", "--event_end_dt"):
        event_end_dt = current_value

if output_file == "":
    print("An output file is required.")
    sys.exit(-2)

if input_file == "":
    print("An input file is required.")
    sys.exit(-3)

# Validate date time values
try:
    if event_start_dt != None:
        event_start_dt = datetime.strptime(event_start_dt, '%Y-%m-%d %H:%M:%S')
    if event_end_dt != None:
        event_end_dt = datetime.strptime(event_end_dt, '%Y-%m-%d %H:%M:%S')
except ValueError:
    print("Invalid date time values")
    sys.exit(-4)

# Make sure event_end_dt is not > event_start_dt
if (event_start_dt != None and event_end_dt != None) and event_start_dt > event_end_dt:
    print("event_start_dt should not be greater than event_end_dt")
    sys.exit(-4)


start(input_file, output_file, event_start_dt, event_end_dt)
